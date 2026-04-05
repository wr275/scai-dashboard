"""
SCAI Dashboard — Server-side Excel Parser
Mirrors the logic in SCAI_Visual_Dashboard.html (parseBaseline / parseWeeklyReport)
"""

import openpyxl
import re
from datetime import datetime, date


# ── Field aliases (mirrors JS FIELD_ALIASES) ─────────────────────────────────
FIELD_ALIASES = {
    'id': ['project id', 'project_id', 'pid', 'id', 'proj id', 'project no', 'project number'],
    'name': ['project name', 'name', 'project title', 'title'],
    'vertical': ['vertical', 'department', 'dept', 'business unit', 'unit', 'stream'],
    'owner': ['owner', 'project owner', 'project manager', 'pm', 'lead', 'project lead', 'manager'],
    'startDate': ['start date', 'start', 'commencement date', 'kickoff date', 'kick-off date'],
    'targetEnd': ['target end date', 'target end', 'end date', 'completion date', 'target completion',
                  'expected end', 'planned end', 'due date'],
    'desc': ['description', 'desc', 'project description', 'summary', 'brief', 'overview'],
    'budget': ['budget', 'total budget', 'project budget', 'approved budget'],
    'phase': ['phase', 'current phase', 'project phase', 'status'],
}

VALID_PHASES = ['Initiation', 'Planning', 'Execution', 'Review & Approval', 'Deployment', 'Closed']
VALID_VERTICALS = [
    'ICT',
    'SC Solutions',
    'Innovation, Partnership & Platforms',
    'Strategy & Planning',
    'AI & Data',
]


def fmt_date(val):
    """Normalise any date value to ISO string YYYY-MM-DD or ''."""
    if val is None:
        return ''
    if isinstance(val, (datetime, date)):
        return val.strftime('%Y-%m-%d')
    s = str(val).strip()
    if not s:
        return ''
    # Try common formats
    for fmt in ('%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y', '%m/%d/%Y', '%d %b %Y', '%d %B %Y', '%Y/%m/%d'):
        try:
            return datetime.strptime(s, fmt).strftime('%Y-%m-%d')
        except ValueError:
            pass
    return s  # return as-is if can't parse


def alias_match(header, field):
    """Return True if header matches any alias for field."""
    h = str(header).strip().lower()
    return h in FIELD_ALIASES.get(field, [])


def find_alias(row_dict, field):
    """Find value in a dict row by alias matching."""
    aliases = FIELD_ALIASES.get(field, [])
    for key, val in row_dict.items():
        if str(key).strip().lower() in aliases:
            return str(val).strip() if val is not None else ''
    return ''


def is_banner_row(row_values):
    """True if row looks like a banner/title row (merged or single spanning text)."""
    non_empty = [v for v in row_values if v is not None and str(v).strip()]
    if len(non_empty) == 1:
        text = str(non_empty[0]).lower()
        if any(kw in text for kw in ['template', 'nmdc', 'scai', 'baseline', 'weekly', 'report', 'tracker']):
            return True
    return False


def sheet_to_list(ws, skip_rows=0):
    """Convert worksheet to list of row dicts, skipping banner rows."""
    rows = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i < skip_rows:
            continue
        rows.append(list(row))
    return rows


# ── Parse Baseline ─────────────────────────────────────────────────────────────
def parse_baseline(wb, pid_hint=''):
    """Parse a Baseline workbook. Returns partial project dict."""
    proj = {
        'id': pid_hint,
        'name': '',
        'vertical': '',
        'owner': '',
        'startDate': '',
        'targetEnd': '',
        'desc': '',
        'budget': '',
        'phase': 'Initiation',
        'milestones': [],
        'plannedProgress': [],
    }

    # ── Project Info sheet ────────────────────────────────────────────────────
    info_ws = wb['Project Info'] if 'Project Info' in wb.sheetnames else None
    if info_ws:
        rows = list(info_ws.iter_rows(values_only=True))

        def find_val(label):
            label_lower = label.lower()
            for row in rows:
                if row is None:
                    continue
                # Template: col A empty, col B = label, col C = value
                if len(row) > 2 and row[1] is not None and str(row[1]).strip().lower() == label_lower:
                    return str(row[2]).strip() if row[2] is not None else ''
                # Fallback: col A = label, col B = value
                if len(row) > 1 and row[0] is not None and str(row[0]).strip().lower() == label_lower:
                    return str(row[1]).strip() if row[1] is not None else ''
            return ''

        proj['id']        = find_val('project id') or pid_hint
        proj['name']      = find_val('project name')
        proj['vertical']  = find_val('vertical')
        proj['owner']     = find_val('owner')
        proj['startDate'] = fmt_date(find_val('start date'))
        proj['targetEnd'] = fmt_date(find_val('target end date'))
        proj['desc']      = find_val('description')
        proj['budget']    = find_val('budget')

        # Normalise vertical
        if proj['vertical']:
            vl = proj['vertical'].lower()
            for v in VALID_VERTICALS:
                if v.lower() == vl or v.lower().startswith(vl[:4]):
                    proj['vertical'] = v
                    break

    # ── Milestones sheet ──────────────────────────────────────────────────────
    ms_ws = wb['Milestones'] if 'Milestones' in wb.sheetnames else None
    if ms_ws:
        all_rows = list(ms_ws.iter_rows(values_only=True))
        # Find the real header row — must have BOTH 'Milestone ID' AND 'Milestone Name' as distinct columns
        header_idx = None
        for i, row in enumerate(all_rows[:8]):
            cells = [str(v).strip().lower() for v in row if v is not None and str(v).strip()]
            has_id   = any(c == 'milestone id' for c in cells)
            has_name = any(c == 'milestone name' for c in cells)
            if has_id and has_name:
                header_idx = i
                break

        if header_idx is not None:
            headers = [str(v).strip() if v is not None else '' for v in all_rows[header_idx]]
            for row in all_rows[header_idx + 1:]:
                row_dict = dict(zip(headers, row))
                ms_id   = str(row_dict.get('Milestone ID', '') or '').strip()
                ms_name = str(row_dict.get('Milestone Name', '') or '').strip()
                if not ms_id or not ms_name:
                    continue
                proj['milestones'].append({
                    'id':         ms_id,
                    'name':       ms_name,
                    'phaseTag':   str(row_dict.get('Phase Tag', 'Execution') or 'Execution').strip(),
                    'targetDate': fmt_date(row_dict.get('Target Date')),
                    'weight':     _safe_float(row_dict.get('Weight'), 1),
                })

    # ── Planned Progress sheet ────────────────────────────────────────────────
    pp_ws = (wb['Planned Progress'] if 'Planned Progress' in wb.sheetnames else
             wb['Planned_Progress'] if 'Planned_Progress' in wb.sheetnames else None)
    if pp_ws:
        all_rows = list(pp_ws.iter_rows(values_only=True))
        for row in all_rows:
            if len(row) >= 2 and row[0] is not None and row[1] is not None:
                d = fmt_date(row[0])
                pct = _safe_float(row[1], None)
                if d and pct is not None:
                    proj['plannedProgress'].append({'date': d, 'pct': pct})

    return proj


# ── Parse Weekly Report ────────────────────────────────────────────────────────
def parse_weekly_report(wb):
    """Parse a WeeklyReport workbook. Returns milestoneTracker + weeklyUpdates."""
    result = {'milestoneTracker': [], 'weeklyUpdates': []}

    # ── Milestone Tracker sheet ───────────────────────────────────────────────
    tracker_ws = wb['Milestone Tracker'] if 'Milestone Tracker' in wb.sheetnames else None
    if tracker_ws:
        all_rows = list(tracker_ws.iter_rows(values_only=True))
        header_idx = _find_strict_header_row(all_rows, ['milestone id'], ['status'])
        if header_idx is not None:
            headers = [str(v).strip() if v is not None else '' for v in all_rows[header_idx]]
            for row in all_rows[header_idx + 1:]:
                row_dict = dict(zip(headers, row))
                ms_id = str(row_dict.get('Milestone ID', '') or '').strip()
                if not ms_id:
                    continue
                result['milestoneTracker'].append({
                    'milestoneId':          ms_id,
                    'status':               str(row_dict.get('Status', 'Not Started') or 'Not Started').strip(),
                    'actualCompletionDate': fmt_date(row_dict.get('Actual Completion Date')),
                })

    # ── Weekly Updates sheet ──────────────────────────────────────────────────
    upd_ws = wb['Weekly Updates'] if 'Weekly Updates' in wb.sheetnames else None
    if upd_ws:
        all_rows = list(upd_ws.iter_rows(values_only=True))
        header_idx = _find_strict_header_row(all_rows, ['week date'], ['accomplishments', 'this week'])
        if header_idx is not None:
            # Get raw header values (preserve newlines for blocker col matching)
            raw_headers = [str(v) if v is not None else '' for v in all_rows[header_idx]]
            for row in all_rows[header_idx + 1:]:
                row_dict = dict(zip(raw_headers, row))
                week_date = fmt_date(_find_in_dict(row_dict, ['week date', 'week', 'date']))
                if not week_date:
                    continue
                # Blocker column — header may contain literal newline
                blocker_val = (
                    row_dict.get('Blockers\n(Y/N)') or
                    row_dict.get('Blockers (Y/N)') or
                    row_dict.get('Blocker?') or
                    row_dict.get('Blockers') or 'N'
                )
                result['weeklyUpdates'].append({
                    'weekDate':      week_date,
                    'thisWeek':      str(_find_in_dict(row_dict, ['this week accomplishments', 'this week', 'accomplishments']) or '').strip(),
                    'hasBlocker':    str(blocker_val).upper().startswith('Y'),
                    'blockerDetail': str(_find_in_dict(row_dict, ['blocker details', 'blocker detail', 'blocker']) or '').strip(),
                    'nextWeek':      str(_find_in_dict(row_dict, ['next week actions', 'next week', 'next actions']) or '').strip(),
                    'ragOverride':   str(_find_in_dict(row_dict, ['rag override', 'rag']) or '').strip() or None,
                    'overrideReason': str(_find_in_dict(row_dict, ['override reason', 'reason']) or '').strip() or None,
                })

    return result


# ── Helpers ───────────────────────────────────────────────────────────────────
def _safe_float(val, default=0):
    try:
        return float(val)
    except (TypeError, ValueError):
        return default


def _find_strict_header_row(rows, required, also_one_of=None, max_scan=8):
    """Find header row that contains ALL required terms AND at least one of also_one_of."""
    for i, row in enumerate(rows[:max_scan]):
        cells = [str(v).strip().lower() for v in row if v is not None and str(v).strip()]
        has_required = all(any(req in c for c in cells) for req in required)
        has_optional = (not also_one_of) or any(any(opt in c for c in cells) for opt in also_one_of)
        if has_required and has_optional:
            return i
    return None

def _find_header_row(rows, keywords, max_scan=6):
    """Find the first row whose non-empty values contain any of the keywords."""
    for i, row in enumerate(rows[:max_scan]):
        non_empty = [str(v).strip().lower() for v in row if v is not None and str(v).strip()]
        if any(any(kw in cell for kw in keywords) for cell in non_empty):
            return i
    return None


def _find_in_dict(d, keys):
    """Find first matching key (case-insensitive) in dict."""
    dl = {k.lower(): v for k, v in d.items()}
    for key in keys:
        if key.lower() in dl:
            return dl[key.lower()]
    return None


# ── Auto-detect and parse any Excel file ─────────────────────────────────────
def parse_any_excel(filepath, pid_hint=''):
    """
    Attempt to parse an Excel file as either Baseline or WeeklyReport.
    Returns (type, data) where type is 'baseline' or 'weekly'.
    """
    wb = openpyxl.load_workbook(filepath, data_only=True)
    sheets_lower = [s.lower() for s in wb.sheetnames]

    is_baseline = 'project info' in sheets_lower and 'milestones' in sheets_lower
    is_weekly   = 'milestone tracker' in sheets_lower or 'weekly updates' in sheets_lower

    if is_baseline:
        return 'baseline', parse_baseline(wb, pid_hint)
    elif is_weekly:
        return 'weekly', parse_weekly_report(wb)
    else:
        return 'unknown', None


def extract_pid_from_filename(filename):
    """Extract project ID from filename like P001_Baseline.xlsx → P001."""
    m = re.match(r'^([A-Za-z0-9-]+)_', filename)
    return m.group(1).upper() if m else ''
